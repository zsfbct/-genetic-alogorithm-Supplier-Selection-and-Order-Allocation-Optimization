#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import openpyxl
from openpyxl import load_workbook
from numpy import *
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
import random
wb = openpyxl.load_workbook('D:\交大\大三下\暑假\python.xlsx');
#print(wb)
sheet=wb['Sheet1']
# 获取所有数据
data_list = list(sheet.rows) 
new_data = []  # 新列表储存数据#
for row in data_list:
    # 每一行的数据存储起来 每循环一次清空
    row_data = []
    # 把 row 里面的数据依次提取出来
    for cell in row:
        # 把值添加到 row_data 列表中
        row_data.append(cell.value)
    # 每循环一次，往 new_data 添加一次数据
    new_data.append(row_data)
#print(new_data)

#父代染色体
pop=[]
SS=[]
SS_1=[None]*2318
def parent_gene(gene):
    num_products=[]
    for i in range(1159):
        num_products.append(int(new_data[i][4])+random.randint(-1,1))#暂时设置成sigma=1
        SS.append(random.randint(0,3))
        i+=1
    for item in num_products:
        SS.append(item)
    return SS
#parent_gene(1)
#print(parent_gene(1))

#pop = np.vstack([parent_gene(1) for _ in range(200)])#把n个小人竖向堆叠
#print(pop)

def get_fitness(SS):
    #总利润
    total = 0
    count = 0 
    profit = -1
    for i in range(len(new_data)):
        while profit < 0 or new_data[i][6+SS[i]]>=1000:
            x=SS[1159+i] #第i+1个料号2021年的采购量 X_1i
            w=np.random.normal(new_data[i][4], 1, 1)[0] #第i+1个料号2021年销量 [均值，sigma,个数]
            m=w-x+new_data[i][5] #D1+D2-X
            y=random.randint(0,3)#在父代染色体的前半段不重要
            if x >= 5000:
                 k=new_data[i][14+y]-new_data[i][10+y]*0.9
            else: k=new_data[i][14+y]-new_data[i][10+y]
            if m >= 5000:
                q=new_data[i][14+y]-new_data[i][10+y]*0.9
            else: q=new_data[i][14+y]-new_data[i][10+y]
            z=stats.norm.pdf(x, new_data[i][4], 1) #概率
            n=new_data[i][18+y] #T 产品开发周期
            o=new_data[i][22+y] #Q 质量约束
            profit=z*(x*n*k-new_data[i][26+y]*n*(x-w)+o*m*q)+(1-z)*(x*n*k+o*new_data[i][5]*q)
            SS[i]=y
            SS[i+1159]=x
        total += profit
        i+=1
    for i in range(len(new_data)):
        if SS[i] == 2:
            count+=1
        i+=1
    #print(total)
    #print(count)
    #print(SS)
    return SS

def get_adjusted_fitness(SS):
    order_nanjing = 0 
    order_A = 0 
    order_B = 0 
    order_C = 0 
    for i in range(len(new_data)):
        if SS[i] == 0:
            order_nanjing += (new_data[i][4]+new_data[i][5])*new_data[i][6]
        if SS[i] == 1:
            order_A += (new_data[i][4]+new_data[i][5])*new_data[i][7]
        if SS[i] == 2:
            order_B += (new_data[i][4]+new_data[i][5])*new_data[i][8]
        if SS[i] == 3:
            order_C += (new_data[i][4]+new_data[i][5])*new_data[i][9]
        i+=1
    while order_nanjing <= 0.1*(order_nanjing+order_A+order_B+order_C) or order_A <= 1.3*10e6 or order_B <= 1.3*10e6 or order_C <= 1.3*10e6:
        get_fitness(SS)
        for i in range(len(new_data)):
            if SS[i] == 0:
                order_nanjing += (new_data[i][4]+new_data[i][5])*new_data[i][6]
            if SS[i] == 1:
                order_A += (new_data[i][4]+new_data[i][5])*new_data[i][7]
            if SS[i] == 2:
                order_B += (new_data[i][4]+new_data[i][5])*new_data[i][8]
            if SS[i] == 3:
                order_C += (new_data[i][4]+new_data[i][5])*new_data[i][9]
            i+=1
    print(SS)
    return(SS)


parent_gene(1)
#get_fitness(SS)
get_adjusted_fitness(SS)


#print(new_data)  
# 保存 提供保存的路径，不是之前的路径那就是另存为
#wb.save(r'D:\交大\大三下\暑假\python.xlsx')
# 关闭
#
wb.close()

 

"""

Visualize Genetic Algorithm to find the shortest path for travel sales problem.



Visit my tutorial website for more: https://morvanzhou.github.io/tutorials/

"""


import numpy as np
import openpyxl
from openpyxl import load_workbook
from numpy import *
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
import random
wb = openpyxl.load_workbook('D:\交大\大三下\暑假\python.xlsx');
sheet=wb['Sheet1']
data_list = list(sheet.rows) 
new_data = []  
for row in data_list:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    new_data.append(row_data)


DNA_SIZE = 2318  # DNA size 染色体长度

CROSS_RATE = 0.15#交叉概率

MUTATE_RATE = 0.02 #变异概率

POP_SIZE = 200 #小人个数

N_GENERATIONS = 500 #繁殖代数



def get_parent(ITEMS):
    pop=[]
    SS=[]
    SS_1=[None]*2318
    num_products=[]
    for i in range(ITEMS):
        num_products.append(int(new_data[i][4])+random.randint(-1,1))#暂时设置成sigma=1
        SS.append(random.randint(0,3))
        i+=1
    for item in num_products:
        SS.append(item)
    profit = -1
    order_nanjing = 0 
    order_A = 0 
    order_B = 0 
    order_C = 0 
    while order_nanjing <= 0.1*(order_nanjing+order_A+order_B+order_C) or order_A <= 1.3*10e6 or order_B <= 1.3*10e6 or order_C <= 1.3*10e6:
        total = 0
        for i in range(len(new_data)):
            while (new_data[i][14+SS[i]]-new_data[i][10+SS[i]])<0 or new_data[i][6+SS[i]]>=1000:
                y=random.randint(0,3)#在父代染色体的前半段不重要
                SS[i]=y
            x=SS[1159+i] #第i+1个料号2021年的采购量 X_1i
            y_1=SS[i]
            w=np.random.normal(new_data[i][4], 1, 1)[0] #第i+1个料号2021年销量 [均值，sigma,个数]
            m=w-x+new_data[i][5] #D1+D2-X
            if x >= 5000:
                 k=new_data[i][14+y_1]-new_data[i][10+y_1]*0.9
            else: k=new_data[i][14+y_1]-new_data[i][10+y_1]
            if m >= 5000:
                q=new_data[i][14+y_1]-new_data[i][10+y_1]*0.9
            else: q=new_data[i][14+y_1]-new_data[i][10+y_1]
            z=stats.norm.pdf(x, new_data[i][4], 1) #概率
            n=new_data[i][18+y_1] #T 产品开发周期
            o=new_data[i][22+y_1] #Q 质量约束
            profit=z*(x*n*k-new_data[i][26+y_1]*n*(x-w)+o*m*q)+(1-z)*(x*n*k+o*new_data[i][5]*q)
            SS[i+1159]=x
            total += profit
            i+=1
        for i in range(len(new_data)):
            if SS[i] == 0:
                order_nanjing += (new_data[i][4]+new_data[i][5])*new_data[i][6]
            if SS[i] == 1:
                order_A += (new_data[i][4]+new_data[i][5])*new_data[i][7]
            if SS[i] == 2:
                order_B += (new_data[i][4]+new_data[i][5])*new_data[i][8]
            if SS[i] == 3:
                order_C += (new_data[i][4]+new_data[i][5])*new_data[i][9]
            i+=1
    #print(SS)
    c=np.stack(SS,axis=0)
    return c



class GA(object): #遗传算法类

    
    def __init__(self, DNA_size, cross_rate, mutation_rate, pop_size, ):

        self.DNA_size = DNA_size

        self.cross_rate = cross_rate

        self.mutate_rate = mutation_rate

        self.pop_size = pop_size

        self.pop = np.vstack([get_parent(1159) for _ in range(pop_size)])#把n个小人竖向堆叠



    def get_fitness(self,SSOA_1):     #SSOA是啥？  
        total=[None]*200
        for j in range(200):
            SSOA = SSOA_1[j]
            total_1=0
            for i in range(len(new_data)):
                x=SSOA[1159+i] #第i+1个料号2021年的采购量 X_1i
                y_1=SSOA[i]
                w=np.random.normal(new_data[i][4], 1, 1)[0] #第i+1个料号2021年销量 [均值，sigma,个数]
                m=w-x+new_data[i][5] #D1+D2-X
                if x >= 5000:
                     k=new_data[i][14+y_1]-new_data[i][10+y_1]*0.9
                else: k=new_data[i][14+y_1]-new_data[i][10+y_1]
                if m >= 5000:
                    q=new_data[i][14+y_1]-new_data[i][10+y_1]*0.9
                else: q=new_data[i][14+y_1]-new_data[i][10+y_1]
                z=stats.norm.pdf(x, new_data[i][4], 1) #概率
                n=new_data[i][18+y_1] #T 产品开发周期
                o=new_data[i][22+y_1] #Q 质量约束
                profit=z*(x*n*k-new_data[i][26+y_1]*n*(x-w)+o*m*q)+(1-z)*(x*n*k+o*new_data[i][5]*q)
                SSOA[i+1159]=x
                total_1 += profit
                i+=1
            total[j] = total_1  
            j+=1
        return total



    def select(self, total):
        sum = 0
        for i in range (self.pop_size):
            sum += total[i]
            i+=1
        idx = np.random.choice(np.arange(self.pop_size), size=self.pop_size, replace=True, p=total / sum)
        #np.arrange()是生成步长为1，总长为self.pop_size的一个从0开始的排列
        #p是被采样的概率
        return self.pop[idx]



    def crossover(self, parent, pop):

        if np.random.rand() < self.cross_rate:             #np.random.rand()是随机生成一个数

            i_ = np.random.randint(0, self.pop_size, size=1)                        # select another individual from pop

            cross_points = np.random.randint(0, 2, 2318).astype(np.bool)   # choose crossover points

            keep_city = parent[cross_points]                                       # find the city number

            swap_city = pop[i_,cross_points ] #找另一个独立的染色体中 选中city的对应位置，没出现的话返回true
            #np.ravel()将多维数组转化为一维数组
            
            parent[cross_points] =  swap_city
            
            pop[i_,cross_points ] =  keep_city 

        return parent



    def mutate(self, child):   #交换两个city的位置

        for point in range(1159):

            if np.random.rand() < self.mutate_rate:

                mutate_point = np.random.randint(0, 1159)+1159
                
                child[ mutate_point]+=np.random.normal(0, 1, 1)[0]

        return child



    def evolve(self, fitness):

        pop = self.select(fitness)

        pop_copy = pop.copy()

        for parent in pop:  # for every parent

            child = self.crossover(parent, pop_copy)

            child = self.mutate(child)

            parent[:] = child

        self.pop = pop





ga = GA(DNA_size=DNA_SIZE, cross_rate=CROSS_RATE, mutation_rate=MUTATE_RATE, pop_size=POP_SIZE)


for generation in range(N_GENERATIONS):


    total = ga.get_fitness(ga.pop)

    ga.evolve(total)

    best_idx = np.argmax(total)

    print('Gen:', generation, '| best fit: %.2f' % total[best_idx],)

